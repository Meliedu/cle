"""Reader for the placement scoring workbook.

The restricted half of the package: per-item metadata and the answer key, plus
the two things v1.3 added -- the ``Item Review`` teacher re-review ledger, and
the operational thresholds and course map from ``Rules & Mapping``.

Reading the rules as data rather than trusting a constant is the point of the
last one. ``import_placement_bank`` compares them against the policy it records
against a version, so a workbook that quietly retunes a threshold cannot be
imported under rules transcribed from an older one. The ``Summary`` formulas
are checked the same way, against the text ``app/services/placement_scoring.py``
was written from.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl

from scripts.placement_sources import FORMS, ITEMS_PER_FORM, ExtractError

#: One row of a sheet read with ``values_only=True``: a tuple of cell values,
#: any of which may be ``None`` for an empty cell.
Row = tuple[Any, ...]

#: The twelve rubric dimensions the teacher scores each item on in the v1.3
#: ``Item Review`` sheet.
_RUBRIC_DIMENSIONS = (
    "NAT", "KEY", "STM", "DST", "CON", "BND", "CUE", "FAIR", "AUD", "TIM", "EXP", "MET",
)


def read_workbook(path: Path) -> dict[str, Any]:
    """Read item metadata, the key matrix, the review ledger, and the rules."""
    workbook = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    try:
        out = {
            "metadata": _read_item_metadata(workbook),
            "key_matrix": _read_key_matrix(workbook),
            "review": _read_item_review(workbook),
            "rules": _read_rules(workbook),
        }
    finally:
        workbook.close()

    # Formulas, not their cached results, so a second pass with data_only off.
    formulas = openpyxl.load_workbook(str(path), data_only=False, read_only=True)
    try:
        _assert_scoring_formulas(formulas)
    finally:
        formulas.close()
    return out


#: ``Rules & Mapping`` rule label -> the :class:`ScoringPolicy` field it sets.
#: Only the thresholds the app actually applies are mapped; the sheet's other
#: rows are documentation.
_RULE_LABELS = {
    "Items per form": "items_per_form",
    "Items per legacy band": "items_per_band",
    "Band minimum": "band_minimum",
    "Cumulative threshold": "cumulative_threshold",
    "Incomplete trigger": "incomplete_answers_below",
    "Fast-duration trigger": "fast_duration_minutes",
    "Long-duration trigger": "long_duration_minutes",
    "Attempt spread trigger": "attempt_spread_bands",
}

#: ``Summary`` cells whose formula *is* the scoring rule, with the formula the
#: app was written against. Whitespace is normalised before comparing; nothing
#: else is. A package that changes any of these has changed how a learner is
#: placed, and must not be imported against rules transcribed from an older
#: workbook -- so a mismatch stops extraction rather than raising a warning.
_SCORING_FORMULAS = {
    "highest_sustained_band": (
        '=IF(AND(B16>=3,SUM(B11:B16)/30>=0.7),6,IF(AND(B15>=3,SUM(B11:B15)/25>=0.7),5,'
        'IF(AND(B14>=3,SUM(B11:B14)/20>=0.7),4,IF(AND(B13>=3,SUM(B11:B13)/15>=0.7),3,'
        'IF(AND(B12>=3,SUM(B11:B12)/10>=0.7),2,IF(AND(B11>=3,B11/5>=0.7),1,0))))))'
    ),
    "provisional_course": (
        '=IF(B19<=1,"LANG1511",IF(B19=2,"LANG1512",IF(B19=3,"LANG1513",'
        'IF(B19=4,"LANG1514","LANG1515"))))'
    ),
    "lower_band_break": (
        '=IF(B19<=1,FALSE,IF(B19=2,B11<=2,IF(B19=3,MIN(B11:B12)<=2,'
        'IF(B19=4,MIN(B11:B13)<=2,IF(B19=5,MIN(B11:B14)<=2,MIN(B11:B15)<=2)))))'
    ),
    "clear_next_band_boundary": (
        '=IF(B19=0,FALSE,IF(B19=6,TRUE,IF(B19=1,B12<=2,IF(B19=2,B13<=2,'
        'IF(B19=3,B14<=2,IF(B19=4,B15<=2,B16<=2))))))'
    ),
    "confidence": (
        '=IF(OR(B19=0,D6<27,\'Entry\'!F4<12,\'Entry\'!F4>45,B21=TRUE),"Review",'
        'IF(AND(D6=30,B22=TRUE,IF(B19=1,B11,IF(B19=2,B12,IF(B19=3,B13,'
        'IF(B19=4,B14,IF(B19=5,B15,B16)))))>=4),"High","Medium"))'
    ),
}

#: Row label in ``Summary`` column A -> the key in :data:`_SCORING_FORMULAS`.
_SUMMARY_ROWS = {
    "Highest sustained band": "highest_sustained_band",
    "Provisional course": "provisional_course",
    "Lower-band break": "lower_band_break",
    "Clear next-band boundary": "clear_next_band_boundary",
    "Confidence": "confidence",
}


def _read_rules(workbook: openpyxl.Workbook) -> dict[str, Any]:
    """The thresholds and course map the workbook states, as data.

    Read so the values recorded against an imported version can be proved equal
    to the workbook's rather than trusted to a constant somebody remembered to
    update.
    """
    if "Rules & Mapping" not in workbook.sheetnames:
        raise ExtractError("workbook has no 'Rules & Mapping' sheet")
    rows = list(workbook["Rules & Mapping"].iter_rows(values_only=True))

    thresholds: dict[str, Any] = {}
    course_map: dict[str, str] = {}
    for row in rows:
        if not row or row[0] is None or len(row) < 2 or row[1] is None:
            continue
        label = str(row[0]).strip()
        if label in _RULE_LABELS:
            value = row[1]
            thresholds[_RULE_LABELS[label]] = (
                float(value) if isinstance(value, float) else int(value)
            )
        elif label.startswith(("Below Band", "Legacy HSK")):
            course_map[label] = str(row[1]).strip()

    missing = sorted(set(_RULE_LABELS.values()) - set(thresholds))
    if missing:
        raise ExtractError(f"Rules & Mapping is missing thresholds: {missing}")
    if len(course_map) != 7:
        raise ExtractError(
            f"Rules & Mapping: expected 7 course-map rows, read {len(course_map)}"
        )
    return {"thresholds": thresholds, "course_map": course_map}


def _assert_scoring_formulas(workbook: openpyxl.Workbook) -> None:
    """The workbook's scoring formulas must be the ones the app implements."""
    if "Summary" not in workbook.sheetnames:
        raise ExtractError("workbook has no 'Summary' sheet")

    found: dict[str, str] = {}
    for row in workbook["Summary"].iter_rows(values_only=True):
        if not row or row[0] is None or len(row) < 2:
            continue
        name = _SUMMARY_ROWS.get(str(row[0]).strip())
        if name and isinstance(row[1], str):
            found[name] = re.sub(r"\s+", "", row[1])

    for name, expected in _SCORING_FORMULAS.items():
        actual = found.get(name)
        if actual is None:
            raise ExtractError(f"Summary sheet has no {name!r} formula")
        if actual != re.sub(r"\s+", "", expected):
            raise ExtractError(
                f"Summary {name!r} formula changed; placement_scoring.py "
                f"transcribes the old one.\n  workbook: {actual}\n  expected: "
                f"{re.sub(r'[ ]+', '', expected)}"
            )


def _sheet_rows(
    workbook: openpyxl.Workbook, name: str, header_cell: str
) -> tuple[dict[str, int], list[Row]]:
    if name not in workbook.sheetnames:
        raise ExtractError(f"workbook has no {name!r} sheet")
    rows = list(workbook[name].iter_rows(values_only=True))
    header_row = next(
        (i for i, row in enumerate(rows) if row and row[0] == header_cell), None
    )
    if header_row is None:
        raise ExtractError(f"{name} sheet has no {header_cell!r} header row")
    header = [str(c).strip() if c is not None else "" for c in rows[header_row]]
    return {n: i for i, n in enumerate(header) if n}, rows[header_row + 1 :]


def _read_item_metadata(workbook: openpyxl.Workbook) -> dict[tuple[str, int], dict[str, Any]]:
    index, rows = _sheet_rows(workbook, "Item Metadata", "Form")
    required = [
        "Form", "Q", "Item ID", "Slot", "Legacy band", "Skill", "Item type",
        "Response", "Topic", "Subskill", "Expected sec", "Playback", "Key",
    ]
    missing = [name for name in required if name not in index]
    if missing:
        raise ExtractError(f"Item Metadata is missing columns: {missing}")

    out: dict[tuple[str, int], dict[str, Any]] = {}
    for row in rows:
        if not row or index["Form"] >= len(row) or row[index["Form"]] is None:
            continue
        key = (str(row[index["Form"]]).strip(), int(row[index["Q"]]))
        if key in out:
            raise ExtractError(f"Item Metadata has two rows for {key}")
        out[key] = {
            "item_id": str(row[index["Item ID"]]).strip(),
            "slot": str(row[index["Slot"]]).strip(),
            "legacy_band": int(row[index["Legacy band"]]),
            "skill": str(row[index["Skill"]]).strip(),
            "item_type": str(row[index["Item type"]]).strip(),
            "response_format": str(row[index["Response"]]).strip(),
            "topic": _optional(row, index, "Topic"),
            "subskill": _optional(row, index, "Subskill"),
            "expected_seconds": int(row[index["Expected sec"]]),
            "audio_playback": (
                int(row[index["Playback"]]) if row[index["Playback"]] else None
            ),
            "correct_answer": str(row[index["Key"]]).strip(),
            "target_vocabulary": _optional(row, index, "Target vocab"),
            "target_grammar": _optional(row, index, "Target grammar/discourse"),
            "copyright_status": _optional(row, index, "Copyright status"),
            "qa_status": _optional(row, index, "QA status"),
        }
    expected = len(FORMS) * ITEMS_PER_FORM
    if len(out) != expected:
        raise ExtractError(f"Item Metadata: expected {expected} rows, read {len(out)}")
    return out


def _read_key_matrix(workbook: openpyxl.Workbook) -> dict[tuple[str, int], dict[str, Any]]:
    """The workbook's own printed answer key, kept as an independent source."""
    index, rows = _sheet_rows(workbook, "Key Matrix", "Q")
    for form in FORMS:
        if form not in index or f"{form}_ID" not in index:
            raise ExtractError(f"Key Matrix has no column for form {form}")

    out: dict[tuple[str, int], dict[str, Any]] = {}
    for row in rows:
        if not row or row[0] is None:
            continue
        number = int(row[index["Q"]])
        for form in FORMS:
            out[(form, number)] = {
                "correct_answer": str(row[index[form]]).strip(),
                "item_id": str(row[index[f"{form}_ID"]]).strip(),
                "legacy_band": int(row[index["Band"]]),
            }
    expected = len(FORMS) * ITEMS_PER_FORM
    if len(out) != expected:
        raise ExtractError(f"Key Matrix: expected {expected} cells, read {len(out)}")
    return out


def _read_item_review(workbook: openpyxl.Workbook) -> dict[tuple[str, int], dict[str, Any]]:
    """The v1.3 ``Item Review`` ledger: what the teacher is asked to confirm.

    New in v1.3. Carried through so the CLE review screen can show why an item
    is in front of a reviewer, rather than making them open the workbook.
    """
    if "Item Review" not in workbook.sheetnames:
        return {}
    index, rows = _sheet_rows(workbook, "Item Review", "Form")
    out: dict[tuple[str, int], dict[str, Any]] = {}
    for row in rows:
        if not row or index["Form"] >= len(row) or row[index["Form"]] is None:
            continue
        out[(str(row[index["Form"]]).strip(), int(row[index["Q"]]))] = {
            "candidate_status": _optional(row, index, "Candidate status"),
            "teacher_priority": _optional(row, index, "Teacher priority"),
            "rubric_total": _int_or_none(row, index, "Rubric total"),
            "rubric": {
                name.lower(): _int_or_none(row, index, name)
                for name in _RUBRIC_DIMENSIONS
                if name in index
            },
            "evidence_excerpt": _optional(row, index, "Evidence excerpt"),
            "review_question": _optional(row, index, "Teacher review question"),
        }
    return out


def _optional(row: Row, index: dict[str, int], name: str) -> str | None:
    position = index.get(name)
    if position is None or position >= len(row):
        return None
    value = row[position]
    return str(value).strip() if value not in (None, "") else None


def _int_or_none(row: Row, index: dict[str, int], name: str) -> int | None:
    value = _optional(row, index, name)
    try:
        return int(value) if value is not None else None
    except ValueError:
        return None


